import uuid
from io import BytesIO
from typing import List

from fastapi import APIRouter, Depends, Request
from openpyxl import Workbook
from sqlalchemy.orm import Session
from starlette.responses import StreamingResponse

from app.api.feedbacks.crud import (
    get_feedback,
    get_feedback_by_id,
    create_feedback,
    delete_feedback,
    get_answers,
)
from app.api.questions.crud import get_question_for_table
from app.api.schemas import Response, FeedbacksSchema
from app.db import get_db
from app.utils.auth_middleware import get_current_user

router = APIRouter()


@router.get("/download")
async def download_feedbacks(
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    alphabet = [
        "A",
        "B",
        "C",
        "D",
        "E",
        "F",
        "G",
        "H",
        "I",
        "J",
        "K",
        "L",
        "M",
        "N",
        "O",
        "P",
        "Q",
        "R",
        "S",
        "T",
        "U",
        "V",
        "W",
        "X",
        "Y",
        "Z",
        "AA",
        "AB",
        "AC",
        "AD",
        "AE",
        "AF",
        "AG",
        "AH",
        "AI",
        "AJ",
        "AK",
        "AL",
        "AM",
        "AN",
        "AO",
        "AP",
        "AQ",
        "AR",
        "AS",
        "AT",
        "AU",
        "AV",
        "AW",
        "AX",
        "AY",
        "AZ",
        "BA",
        "BB",
        "BC",
        "BD",
        "BE",
        "BF",
        "BG",
        "BH",
        "BI",
        "BJ",
        "BK",
        "BL",
        "BM",
        "BN",
        "BO",
        "BP",
        "BQ",
        "BR",
        "BS",
        "BT",
        "BU",
    ]
    wb = Workbook()
    ws = wb.active
    _feedback = get_feedback(db)
    questions = get_question_for_table(db)

    for index, question in enumerate(questions):
        ws[f"{alphabet[index]}{1}"] = question.name
        ws.column_dimensions[f"{alphabet[index]}"].width = len(question.name)
    ws[f"{alphabet[len(questions)+1]}{1}"] = "Jami"

    feedbacks = [
        {
            "id": feedback.id,
            "keywords": feedback.keywords,
            "answers": [
                {
                    "id": question.id,
                    "question": question.name,
                    "feedback": get_answers(db, feedback.id, question.id),
                }
                for question in questions
            ],
        }
        for feedback in _feedback
    ]
    count = 0
    for index, feedback in enumerate(feedbacks):
        for idx, answer in enumerate(feedback["answers"]):
            if answer["feedback"]:
                ws[f"{alphabet[idx]}{index+2}"] = answer["feedback"][0]
            else:
                ws[f"{alphabet[idx]}{index+2}"] = ""
        feedbacks1 = [
            int(item)
            for sublist in feedback["answers"]
            for item in sublist["feedback"]
            if item.isdigit()
        ]
        count += sum(feedbacks1)
        ws[f'{alphabet[len(feedback["answers"]) + 1]}{index + 2}'] = sum(feedbacks1)

    ws[f"{alphabet[len(questions)]}{len(feedbacks) + 3}"] = "Jami: "
    ws[f"{alphabet[len(questions)+1]}{len(feedbacks) + 3}"] = count
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=result.xlsx"},
    )


@router.get("/v2")
async def get_feedbacks_route(
    req: Request,
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    _feedbacks = get_feedback(db, search=req.query_params.get("search"))
    _questions = get_question_for_table(db)
    return Response(
        code=200,
        status="ok",
        message="success",
        result=[
            {
                "id": feedback.id,
                "keywords": feedback.keywords,
                "answers": [
                    {
                        "id": question.id,
                        "question": question.name,
                        "feedback": get_answers(db, feedback.id, question.id),
                    }
                    for question in _questions
                ],
            }
            for feedback in _feedbacks
        ],
        total=0,
    ).dict()


@router.get("/{feedback_id}")
async def get_feedback_by_id_route(
    feedback_id: uuid.UUID,
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    _feedback = get_feedback_by_id(db, feedback_id)
    return Response(
        code=200, status="ok", message="success", result=_feedback
    ).model_dump()


@router.get("/")
async def get_feedbacks_route(
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    _feedbacks = get_feedback(db)

    return Response(
        code=200,
        status="ok",
        message="success",
        result=[
            {
                "id": feedback.id,
                "feedback": feedback.feedback,
                "question": feedback.question,
                "created_at": feedback.created_at,
                "updated_at": feedback.updated_at,
            }
            for feedback in _feedbacks
        ],
        total=0,
    ).dict()


@router.post("/")
async def create_feedback_route(
    feedback: List[FeedbacksSchema], db: Session = Depends(get_db)
):
    create_feedback(db, feedback)
    return Response(code=201, status="ok", message="created").dict()


@router.delete("/{feedback_id}")
async def delete_feedback_route(
    feedback_id: uuid.UUID,
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    delete_feedback(db, feedback_id)
    return Response(
        code=200,
        status="ok",
        message="deleted",
    ).model_dump()
