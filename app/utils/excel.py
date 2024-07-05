from openpyxl import Workbook


wb = Workbook()
ws = wb.active

ws["A1"] = "Login"
ws["B1"] = "Soni"
ws["C1"] = "Jami summa"

for index, item in enumerate():
    ws[f"A{index + 2}"] = item["login"]
    ws[f"B{index + 2}"] = item["total"]
    ws[f"C{index + 2}"] = item["sum"]

ws.column_dimensions["A"].width = 30
ws.column_dimensions["c"].width = 30

wb.save("result.xlsx")
